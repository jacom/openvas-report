from io import BytesIO
from django.http import HttpResponse
from django.db.models import Count, Q
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import PieChart, Reference
from openpyxl.utils import get_column_letter
from scanner.models import Vulnerability, AffectedHost


SEVERITY_FILLS = {
    'Critical': PatternFill(start_color='DC3545', end_color='DC3545', fill_type='solid'),
    'High': PatternFill(start_color='FD7E14', end_color='FD7E14', fill_type='solid'),
    'Medium': PatternFill(start_color='FFC107', end_color='FFC107', fill_type='solid'),
    'Low': PatternFill(start_color='0DCAF0', end_color='0DCAF0', fill_type='solid'),
    'Info': PatternFill(start_color='6C757D', end_color='6C757D', fill_type='solid'),
}

HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
HEADER_FILL = PatternFill(start_color='212529', end_color='212529', fill_type='solid')
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin'),
)


def _style_header_row(ws, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN_BORDER


def _auto_column_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)


def generate_excel(report):
    wb = Workbook()

    # Sheet 1: Raw vulnerability data
    ws1 = wb.active
    ws1.title = 'Vulnerabilities'
    headers = ['Severity', 'CVSS', 'Vulnerability', 'Family', 'Host IP',
               'Hostname', 'Port', 'Protocol', 'CVEs', 'Description', 'Solution']
    ws1.append(headers)
    _style_header_row(ws1, len(headers))

    vulns = Vulnerability.objects.filter(report=report).prefetch_related('affected_hosts')
    row_num = 2
    for vuln in vulns:
        for host in vuln.affected_hosts.all():
            ws1.append([
                vuln.severity,
                vuln.cvss_score,
                vuln.name,
                vuln.family,
                host.ip,
                host.hostname,
                host.port,
                host.protocol,
                ', '.join(vuln.cve_list),
                vuln.description[:1000],
                vuln.solution[:1000],
            ])
            # Color severity cell
            sev_cell = ws1.cell(row=row_num, column=1)
            if vuln.severity in SEVERITY_FILLS:
                sev_cell.fill = SEVERITY_FILLS[vuln.severity]
                if vuln.severity in ('Critical', 'High', 'Info'):
                    sev_cell.font = Font(color='FFFFFF', bold=True)
            row_num += 1

    _auto_column_width(ws1)

    # Sheet 2: Pivot-ready summary (severity x host)
    ws2 = wb.create_sheet('Host Summary')
    ws2.append(['Host IP', 'Critical', 'High', 'Medium', 'Low', 'Info', 'Total'])
    _style_header_row(ws2, 7)

    host_data = AffectedHost.objects.filter(
        vulnerability__report=report
    ).values('ip').annotate(
        critical=Count('id', filter=Q(vulnerability__severity='Critical')),
        high=Count('id', filter=Q(vulnerability__severity='High')),
        medium=Count('id', filter=Q(vulnerability__severity='Medium')),
        low=Count('id', filter=Q(vulnerability__severity='Low')),
        info=Count('id', filter=Q(vulnerability__severity='Info')),
        total=Count('id'),
    ).order_by('-critical', '-high', '-total')

    for h in host_data:
        ws2.append([h['ip'], h['critical'], h['high'], h['medium'], h['low'], h['info'], h['total']])

    _auto_column_width(ws2)

    # Sheet 3: Severity distribution chart
    ws3 = wb.create_sheet('Charts')
    ws3.append(['Severity', 'Count'])
    ws3.append(['Critical', report.critical_count])
    ws3.append(['High', report.high_count])
    ws3.append(['Medium', report.medium_count])
    ws3.append(['Low', report.low_count])
    ws3.append(['Info', report.info_count])

    pie = PieChart()
    pie.title = 'Severity Distribution'
    pie.style = 10
    labels = Reference(ws3, min_col=1, min_row=2, max_row=6)
    data = Reference(ws3, min_col=2, min_row=1, max_row=6)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.width = 18
    pie.height = 14
    ws3.add_chart(pie, 'D2')

    _auto_column_width(ws3)

    # Write response
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="openvas_report_{report.id}.xlsx"'
    return response
